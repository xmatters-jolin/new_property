"""Queries for and processes xmatters events

.. _Google Python Style Guide:
   http://google.github.io/styleguide/pyguide.html

"""

import json
import sys
import pprint
from io import TextIOBase
import urllib.parse

import requests
from requests.auth import HTTPBasicAuth
# Import `load_workbook` module from `openpyxl`
from openpyxl import Workbook, load_workbook
# Import relevant modules from `openpyxl.utils`
from openpyxl.utils import get_column_letter, column_index_from_string

import config
import np_logger

_logger = None

def _log_xm_error(url, response):
    """Captures and logs errors
        
        Logs the error caused by attempting to call url.
        
        Args:
        url (str): The location being requested that caused the error
        response (object): JSON object that holds the error response
        """
    body = response.json()
    if response.status_code == 404:
        _logger.warn(config.ERR_INITIAL_REQUEST_FAILED_MSG,
                     response.status_code, url)
    else:
        _logger.error(config.ERR_INITIAL_REQUEST_FAILED_MSG,
                      response.status_code, url)
        _logger.error('Response - code: %s, reason: %s, message: %s',
                    str(body['code']) if 'code' in body else "none",
                    str(body['reason']) if 'reason' in body else "none",
                    str(body['message']) if 'message' in body else "none")

def _match_field(not_matching_values, cell, prop_obj, field_name):
    """Compares source with existing object
        
        Compares cell to object, and if not matching puts it in the array
        
        Args:
        not_matching_values: List of not matching values that is updated
        cell: The current cell
        site_obj: The current Site object
        field_name: Name of field in site_obj
        """
    if cell.value != prop_obj[field_name]:
        not_matching_values.append('%s:(cell=[%s],object=[%s])' % (field_name, cell.value, prop_obj[field_name]))

def _add_site(site_name, row):
    """Attempst to add a new Site object based on the Cell.
        
        Creates a dict object to pass to xMatters to create a new site
        
        Args:
        site_name: The name of the site to add
        row: The row containing data to add
        """
    _logger.debug("Attempting to add Site: %s", site_name)
    
    # Setup object to post
    data = {}
    for cell in row:
        if cell.column == 'D': # name
            data['name'] = cell.value
        elif cell.column == 'E': # address1
            data['address1'] = cell.value
        elif cell.column == 'F' and cell.value is not None and len(cell.value) > 0: # address2
            data['address2'] = cell.value
        elif cell.column == 'G': # city
            data['city'] = cell.value
        elif cell.column == 'H': # country
            data['country'] = cell.value
        elif cell.column == 'I': # language
            data['language'] = cell.value
        elif cell.column == 'J': # postalCode
            data['postalCode'] = cell.value
        elif cell.column == 'K': # state
            data['state'] = cell.value
        elif cell.column == 'L': # timezone
            data['timezone'] = cell.value
        elif cell.column == 'M' and cell.value is not None and len(cell.value) > 0: # latitude
            data['latitude'] = cell.value
        elif cell.column == 'N' and cell.value is not None and len(cell.value) > 0: # longitude
            data['longitude'] = cell.value
    
    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/sites'
    _logger.debug('Attempting to create site "%s" via url: %s', site_name, url)

    # Initialize loop with first request
    try:
        response = requests.post(url,
                                 headers = {'Content-Type': 'application/json'},
                                 data = json.dumps(data),
                                 auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None

    # If the initial response fails, log and return null
    if response.status_code != 201:
        _log_xm_error(url, response)
        return None

    # Process the response
    site_obj = response.json()
    _logger.info('Created Site "%s" - Id: %s', site_name, site_obj['id'])
    # _logger.info('Created Site "%s" - json body: %s', site_name, pprint.pformat(site_obj))
    return site_obj

def _site_exists(site_name: str):
    """Attempst to retrieve site by name.
        
        If the named site exists, retrieve and return the object.
        If not, return null
        
        Args:
        sites_name (str): Name of site to retrieve
        """
    _logger.debug("Attempting to retrieve Site: %s", site_name)
    
    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/sites/' + urllib.parse.quote(site_name);
    _logger.debug('Attempting to retrieve site "%s" via url: %s', site_name, url)
    
    # Initialize loop with first request
    try:
        response = requests.get(url, auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None
    
    # If the initial response fails, log and return null
    if response.status_code != 200:
        _log_xm_error(url, response)
        return None
    
    # Process the response
    site_obj = response.json()
    # _logger.debug('Site "%s" - json body: %s', str, pprint.pformat(site_obj))
    return site_obj

def _sites_match(row: tuple, site_obj: dict):
    """Compares source with existing object and returns true if they match.
        
        Looks at the details of the retrieved site object and returns false
        if the column values don't match
        
        Args:
        sites_file (Workbook): Open sites worksheet
        """
    _logger.debug('Comparing worksheet with xMatters for site "%s".', site_obj['name'])
    match = True
    not_matching_values = [];
    for cell in row:
        if config.non_prod and cell.column == 'C':
            _match_field(not_matching_values, cell, site_obj, 'id')
        elif not config.non_prod and cell.column == 'B':
            _match_field(not_matching_values, cell, site_obj, 'id')
        elif cell.column == 'E': # address1
            _match_field(not_matching_values, cell, site_obj, 'address1')
        elif cell.column == 'F' and 'address2' in site_obj and len(site_obj['address2']) > 0: # address2
            _match_field(not_matching_values, cell, site_obj, 'address2')
        elif cell.column == 'G': # city
            _match_field(not_matching_values, cell, site_obj, 'city')
        elif cell.column == 'H': # country
            _match_field(not_matching_values, cell, site_obj, 'country')
        elif cell.column == 'I': # language
            _match_field(not_matching_values, cell, site_obj, 'language')
        elif cell.column == 'J': # postalCode
            _match_field(not_matching_values, cell, site_obj, 'postalCode')
        elif cell.column == 'K': # state
            _match_field(not_matching_values, cell, site_obj, 'state')
        elif cell.column == 'L': # timezone
            _match_field(not_matching_values, cell, site_obj, 'timezone')
    if len(not_matching_values) > 0:
        match = False
        _logger.error('Site "%s" DOES NOT MATCH the source worksheet.%s', site_obj['name'],
                 ', '.join(not_matching_values))
    else:
        _logger.info('Site "%s" matches the source worksheet', site_obj['name'])

    return match

def _process_sites(sites_file: Workbook):
    """Retrieves and processes xMatters Site objects.
        
        Retrieves the Sites from the excel sheet and then attempst to verify
        them in xMatters.  If they don't exist yet, then create them.
        Update the spreadsheet with IDs.
        
        Args:
        sites_file (Workbook): Open sites worksheet
        """
    _logger.info('Processing worksheet for Sites.')

    # Get the Sites worksheet
    sites_sheet = sites_file["Sites"];
    
    for row in sites_sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column == 'D':
                _logger.debug('Found cell=%s', cell)
                site_name = cell.value
                site_obj = _site_exists(site_name)
                if site_obj:
                    _logger.info('Processing Site "%s", id=[%s] in the %s environment',
                                 site_name, site_obj['id'],
                                 'Non-Production' if config.non_prod else 'Production')
                    if not _sites_match(row, site_obj):
                        # Update the spreadsheet
                        if config.non_prod:
                            sites_sheet.cell(row=cell.row, column=column_index_from_string('C')).value = site_obj['id']
                        else:
                            sites_sheet.cell(row=cell.row, column=column_index_from_string('B')).value = site_obj['id']
                        sites_sheet.cell(row=cell.row, column=column_index_from_string('M')).value = repr(site_obj['latitude'])
                        sites_sheet.cell(row=cell.row, column=column_index_from_string('N')).value = repr(site_obj['longitude'])
                else:
                    _logger.info('Site "%s" does not exist in the %s environment; adding.',
                                 site_name,
                                 'Non-Production' if config.non_prod else 'Production')
                    site_obj = _add_site(site_name, row)
                    if site_obj:
                        if config.non_prod:
                            sites_sheet.cell(row=cell.row, column=column_index_from_string('C')).value = site_obj['id']
                        else:
                            sites_sheet.cell(row=cell.row, column=column_index_from_string('B')).value = site_obj['id']

def _get_site_id(sites_sheet, site_name: str):
    """Return the UUID of the named site
        
        Looks through the sites sheet for that site row
        
        Args:
        sites_sheet: Worksheet with Site details
        site_name (str): Name of the site to find
        """
    _logger.debug('Searching for site "%s".', site_name)
    found = False
    for row in sites_sheet.iter_rows(min_row=2):
        for cell in row:
            id = ''
            if cell.column == 'D':
                if cell.value == site_name:
                    if config.non_prod:
                        id = sites_sheet.cell(row=cell.row, column=column_index_from_string('C')).value
                        _logger.debug('Found site "%s" with Non-Prod ID of [%s].', site_name, id)
                        return id
                    else:
                        id = sites_sheet.cell(row=cell.row, column=column_index_from_string('B')).value
                        _logger.debug('Found site "%s" with Production ID of [%s].', site_name, id)
                        return id
    return None

def _add_email_device(owner_name: str, owner_id: str, email_addr: str):
    """Attempst to add an email device to a User.
        
        Adds an email device to the user.
        
        Args:
        owner_name: User's targetName
        owner_id: The id of the user to add a device to
        email_addr: The email address to add
        """
    _logger.debug("Attempting to add device to User: %s", owner_name)
    
    # Setup object to post
    data = {
        'name' : config.device_name,
        'owner' : owner_id,
        'deviceType' : config.device_type,
        'recipientType' : 'DEVICE',
        'defaultDevice' : True,
        'emailAddress' : email_addr
    }
    
    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/devices'
    _logger.debug('Attempting to add email device for %s to user "%s" via url: %s\njson body: %s',
                  email_addr,
                  owner_name,
                  url,
                  json.dumps(data))

    # Initialize loop with first request
    try:
        response = requests.post(url,
                                 headers = {'Content-Type': 'application/json'},
                                 data = json.dumps(data),
                                 auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None

    # If the initial response fails, log and return null
    if response.status_code != 201:
        _log_xm_error(url, response)
        return None
        
    # Process the response
    dev_obj = response.json()
    _logger.info('Created Email Device "%s" for "%s" - idy: %s', email_addr, owner_name, dev_obj['id'])
    # _logger.debug('Created Email Device "%s" for "%s" - json body: %s', email_addrowner_name, pprint.pformat(user_obj))
    return dev_obj

def _add_user(target_name, site_id, row):
    """Attempst to add a new User object based on the Cell.
        
        Creates a dict object to pass to xMatters to create a new site
        
        Args:
        target_name: The key of the user to add
        site_id: The id for the related site object
        row: The row containing data to add
        """
    _logger.debug("Attempting to add User: %s", target_name)
    
    # Setup object to post
    data = {
        'targetName' : target_name,
        'site' : site_id,
        'supervisors' : config.supervisor_ids
    }
    for cell in row:
        if cell.column == 'A': # name
            properties = {}
            properties[config.udf_name] = cell.value
            data['properties'] = properties
        elif cell.column == 'E': # firstName
            data['firstName'] = cell.value
        elif cell.column == 'F': # lastName
            data['lastName'] = cell.value
        elif cell.column == 'G': # roles
            in_roles = cell.value.split('|')
            data['roles'] = in_roles

    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/people'
    _logger.debug('Attempting to create user "%s" via url: %s\njson body: %s',
                  target_name,
                  url,
                  json.dumps(data))

    # Initialize loop with first request
    try:
        response = requests.post(url,
                                 headers = {'Content-Type': 'application/json'},
                                 data = json.dumps(data),
                                 auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None

    # If the initial response fails, log and return null
    if response.status_code != 201:
        _log_xm_error(url, response)
        return None
    
    # Process the response
    user_obj = response.json()
    _logger.info('Created User "%s" - idy: %s', target_name, user_obj['id'])
    # _logger.debug('Created User "%s" - json body: %s', target_name, pprint.pformat(user_obj))
    return user_obj

def _get_user(target_name: str):
    """Attempst to retrieve User by targetName.
        
        If the named User exists, retrieve and return the object.
        If not, return null
        
        Args:
        target_name (str): Target Name of User to retrieve
        """
    _logger.debug("Retrieving User: %s", target_name)
    
    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/people/' + urllib.parse.quote(target_name) + '?embed=roles,supervisors';
    _logger.debug('Attempting to retrieve User "%s" via url: %s', target_name, url)
    
    # Initialize loop with first request
    try:
        response = requests.get(url, auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None
    
    # If the initial response fails, log and return null
    if response.status_code != 200:
        _log_xm_error(url, response)
        return None
    
    # Process the response
    user_obj = response.json()
    # _logger.debug('Found User "%s" - json body: %s', target_name, pprint.pformat(user_obj))
    _logger.debug('Found User "%s" - json body.id: %s', target_name, user_obj['id'])
    return user_obj

def _users_match(row: tuple, user_obj: dict):
    """Compares source with existing object and returns true if they match.
        
        Looks at the details of the retrieved User object and returns false
        if the column values don't match
        
        Args:
        row (tuple): The row representing this Admin User from the spreadsheet
        user_obj (dict): The retrieved user
        """
    _logger.debug('Comparing worksheet with xMatters for User "%s".', user_obj['targetName'])
    match = True
    not_matching_values = [];
    for cell in row:
        if cell.column == 'A':
            _match_field(not_matching_values, cell, user_obj['properties'], config.udf_name)
        elif not config.non_prod and cell.column == 'B':
            _match_field(not_matching_values, cell, user_obj, 'id')
        elif config.non_prod and cell.column == 'C':
            _match_field(not_matching_values, cell, user_obj, 'id')
        elif cell.column == 'E': # firstName
            _match_field(not_matching_values, cell, user_obj, 'firstName')
        elif cell.column == 'F': # lastName
            _match_field(not_matching_values, cell, user_obj, 'lastName')
        elif cell.column == 'G': # roles
            ws_roles = cell.value.split('|')
            num_ws_roles = len(ws_roles)
            _logger.debug('Found %d ws_roles.',num_ws_roles)
            ws_role_cnt = 0
            for ws_role in ws_roles:
                _logger.debug('ws_role: %s', ws_role)
                for user_role in user_obj['roles']['data']:
                    _logger.debug("user_role['name']: %s", user_role['name'])
                    if ws_role == user_role['name']:
                        _logger.debug('Matched %s',ws_role)
                        ws_role_cnt += 1
                        break
            if ws_role_cnt != num_ws_roles:
                not_matching_values.append('%s:(cell=[%s],site=[%s])' % ('roles', cell.value, user_obj['roles']['data']))
    if len(not_matching_values) > 0:
        match = False
        _logger.error('User "%s" DOES NOT MATCH the source worksheet.%s', user_obj['targetName'],
                      ', '.join(not_matching_values))
    else:
        _logger.info('User "%s" matches the source worksheet', user_obj['targetName'])
    
    return match

def _process_admins(sites_file: Workbook):
    """Retrieves and processes xMatters Admin User objects.
        
        Retrieves the Admins from the excel sheet and then attempst to verify
        them in xMatters.  If they don't exist yet, then create them.
        Update the spreadsheet with IDs.
        
        Args:
        sites_file (Workbook): Open sites worksheet
        """
    _logger.info('Processing worksheet for Admins.')
    
    # Get the Sites worksheet
    sites_sheet = sites_file["Sites"];

    # Get the Admins worksheet
    admins_sheet = sites_file["Admins"];
    
    # Resolve the default supervisor
    for supervisor in config.supervisors:
        supervisor_obj = _get_user(supervisor)
        if supervisor is None:
            _logger.error('Unable to find default supervisor %s', supervisor)
        else:
            config.supervisor_ids.append(supervisor_obj['id'])

    for row in admins_sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column == 'D':
                _logger.debug('Found cell=%s', cell)
                target_name = cell.value
                site_name = admins_sheet.cell(row=cell.row, column=column_index_from_string('H')).value
                site_id = _get_site_id(sites_sheet, site_name)
                if site_id:
                    user_obj = _get_user(target_name)
                    if user_obj:
                        _logger.info('Processing User "%s", id=[%s] in the %s environment',
                                     target_name, user_obj['id'],
                                     'Non-Production' if config.non_prod else 'Production')
                        if not _users_match(row, user_obj):
                            # Update the spreadsheet
                            if config.non_prod:
                                admins_sheet.cell(row=cell.row, column=column_index_from_string('C')).value = user_obj['id']
                            else:
                                admins_sheet.cell(row=cell.row, column=column_index_from_string('B')).value = user_obj['id']
                    else:
                        _logger.info('Processing User "%s" does not exist in the %s environment; adding.',
                                     target_name,
                                     'Non-Production' if config.non_prod else 'Production')
                        user_obj = _add_user(target_name, site_id, row)
                        if user_obj:
                            _add_email_device(target_name,
                                              user_obj['id'],
                                              admins_sheet.cell(row=cell.row, column=column_index_from_string('I')).value)
                            if config.non_prod:
                                admins_sheet.cell(row=cell.row, column=column_index_from_string('C')).value = user_obj['id']
                            else:
                                admins_sheet.cell(row=cell.row, column=column_index_from_string('B')).value = user_obj['id']
                else:
                    _logger.error('Unable to find Site "%s" for user %s.', site_name, target_name)

def _get_site_id_from_sites_sheet(sites_sheet, site_name):
    """Finds and returns the id for the named site
        
        Iterates through the rows in he sites sheet looking for a matching site name.
        Returns the ID if found, or None.
        
        Args:
        sites_sheet: An active Sites worksheet
        site_name: The name of the site to find
        """
    for col in sites_sheet.iter_cols(min_row=2, min_col=column_index_from_string('D'), max_col=column_index_from_string('D')):
        for cell in col:
            if cell.value == site_name:
                if config.non_prod:
                    return sites_sheet.cell(row=cell.row, column=column_index_from_string('C')).value
                else:
                    return sites_sheet.cell(row=cell.row, column=column_index_from_string('B')).value
    return None

def _get_supervisors_from_admins_sheet(admins_sheet, site_name):
    """Finds and returns the list of admin ids for the named site
        
        Iterates through the rows in the admins sheet looking for a matching site name.
        Returns the ID if found, or None.
        
        Args:
        admins_sheet: An active Admins worksheet
        site_name: The name of the site to find
        """
    supervisors = []
    for col in admins_sheet.iter_cols(min_row=2, min_col=column_index_from_string('H'), max_col=column_index_from_string('H')):
        for cell in col:
            if cell.value == site_name:
                if config.non_prod:
                    supervisors.append(admins_sheet.cell(row=cell.row, column=column_index_from_string('C')).value)
                else:
                    supervisors.append(admins_sheet.cell(row=cell.row, column=column_index_from_string('B')).value)
    return supervisors

def _add_group_members(target_name, site_id, supervisors):
    """Attempst to add supervisors members to Group
        
        Adds supervisors as members to the group
        
        Args:
        target_name: The key of the Group to add
        site_id: The id for the related site object
        supervisors: list of supervisor ids
        """
    _logger.debug("Attempting to add members to Group: %s", target_name)
    members = []
    
    # Add Supervisors as members in the Roster
    for supervisor in supervisors:
        # Setup object to post
        recip = {
            'recipientType' : 'PERSON',
            'id' : supervisor
        }
        data = {
            'recipient' : recip
        }
        
        # Set our resource URLs
        url = config.xmod_url + '/api/xm/1/groups/' + urllib.parse.quote(target_name) + '/shifts/Default%20Shift/members'
        _logger.debug('Attempting to add Supervisor with id[%s] to Group "%s" via url: %s\njson body: %s',
                      supervisor,
                      target_name,
                      url,
                      json.dumps(data))
        
        # Initialize loop with first request
        try:
            response = requests.post(url,
                                     headers = {'Content-Type': 'application/json'},
                                     data = json.dumps(data),
                                     auth=config.basic_auth)
        except requests.exceptions.RequestException as e:
            _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))

        # If the initial response fails, log
        if response.status_code != 200:
            _log_xm_error(url, response)
        else:
            # Process the response
            member_obj = response.json()
            members.append(member_obj['recipient']['id'])
            _logger.info('Added member to Group "%s" - id: %s', target_name, members[len(members)-1])
            # _logger.debug('Added member to Group "%s" - json body: %s', target_name, pprint.pformat(member_obj))

    return members

def _add_group(target_name, site_id, site_name, supervisors):
    """Attempst to add a new Group object
        
        Creates a dict object to pass to xMatters to create a new Group
        
        Args:
        target_name: The key of the user to add
        site_id: The id for the related site object
        site_name: Literal name of the site
        supervisors: list of supervisor ids
        """
    _logger.debug("Attempting to add Group: %s", target_name)
    
    # Setup object to post
    data = {
        'targetName' : target_name,
        'description' : target_name[3:] + ' Administrators',
        'site' : site_id,
        'supervisors' : supervisors,
        'observedByAll' : False
    }

    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/groups'
    _logger.debug('Attempting to create Group "%s" via url: %s\njson body: %s',
                  target_name,
                  url,
                  json.dumps(data))

    # Initialize loop with first request
    try:
        response = requests.post(url,
                                 headers = {'Content-Type': 'application/json'},
                                 data = json.dumps(data),
                                 auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None

    # If the initial response fails, log and return null
    if response.status_code != 201:
        _log_xm_error(url, response)
        return None
        
    # Process the response
    group_obj = response.json()
    _logger.info('Created Group "%s" - id: %s', target_name, group_obj['id'])
    # _logger.debug('Created Group "%s" - json body: %s', target_name, pprint.pformat(group_obj))
        
    return group_obj

def _get_group(target_name: str):
    """Attempst to retrieve Group by targetName.
        
        If the named Group exists, retrieve and return the object.
        If not, return null
        
        Args:
        target_name (str): Target Name of Group to retrieve
        """
    _logger.debug("Attempting to retrieve Group: %s", target_name)
    
    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/groups/' + urllib.parse.quote(target_name) + '?embed=supervisors';
    _logger.debug('Attempting to retrieve Group "%s" via url: %s', target_name, url)
    
    # Initialize loop with first request
    try:
        response = requests.get(url, auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return None
    
    # If the initial response fails, log and return null
    if response.status_code != 200:
        _log_xm_error(url, response)
        return None
    
    # Process the response
    group_obj = response.json()
    # _logger.debug('Found Group "%s" - json body: %s', target_name, pprint.pformat(group_obj))
    _logger.debug('Found Group "%s" - json body.id: %s', target_name, group_obj['id'])
    return group_obj

def _get_group_members(target_name: str, id: str):
    """Attempst to retrieve Group Roster by targetName.
        
        If the named Group exists, retrieve and return the roster.
        If not, return null
        
        Args:
        target_name (str): Target Name of Group to retrieve members of
        id (str): The ID of the group
        """
    _logger.debug("Attempting to retrieve User: %s", target_name)
    group_members = []
    
    # Set our resource URLs
    url = config.xmod_url + '/api/xm/1/groups/' + id + '/members';
    _logger.debug('Attempting to retrieve members of Group "%s" via url: %s', target_name, url)
    
    # Initialize loop with first request
    try:
        response = requests.get(url, auth=config.basic_auth)
    except requests.exceptions.RequestException as e:
        _logger.error(config.ERR_REQUEST_EXCEPTION_CODE, url, repr(e))
        return group_members
    
    # If the initial response fails, log and return null
    if response.status_code != 200:
        _log_xm_error(url, response)
        return group_members
    
    # Process the response
    group_member_obj = response.json()
    # _logger.debug('Found Group "%s" - json body: %s', target_name, pprint.pformat(group_obj))
    if group_member_obj['total'] > 0:
        _logger.debug('Found %d members of Group "%s" - json body.id: %s',
                      group_member_obj['count'], target_name,
                      group_member_obj['data'][0]['group']['id'])
        for member in group_member_obj['data']:
            group_members.append(member['member']['id'])
    else:
        _logger.debug('Found 0 members of Group "%s"', target_name)

    return group_members

def _collect_supervisors(group_obj: dict):
    supervisors = []
    for sup in group_obj['supervisors']['data']:
        supervisors.append(sup['id'])
    return supervisors

def _group_match(groups_sheet, row, site_id: str, supervisors: list, group_obj: dict):
    """Compares source with existing object and returns true if they match.
        
        Looks at the details of the retrieved Group object and returns false
        if the column values don't match
        
        Args:
        groups_sheet: The Groups spreadsheet
        row: Row
        site_id (str): GUID for the related Site
        supervisors (list): Array of supervisor IDs
        group_obj (dict): The retrieved group
        """
    _logger.debug('Comparing worksheet with xMatters for Group "%s".', group_obj['targetName'])
    match = True
    not_matching_values = [];

    if not config.non_prod:
        _match_field(not_matching_values, groups_sheet.cell(row=row, column=column_index_from_string('B')), group_obj, 'id')
    else:
        _match_field(not_matching_values, groups_sheet.cell(row=row, column=column_index_from_string('C')), group_obj, 'id')

    if group_obj['site']['id'] != site_id:
        not_matching_values.append('site:(cell=[%s],group=[%s])' % (site_id, group_obj['site']['id']))

    grp_supervisors = _collect_supervisors(group_obj)
    if len(grp_supervisors) == 0 or not set(supervisors).issubset(set(grp_supervisors)):
        not_matching_values.append('supervisors:(cell=[%s],group=[%s])' % (supervisors, grp_supervisors))

    if group_obj['observedByAll'] != False:
        not_matching_values.append('observedByAll:(cell=[False],group=[%s])' % (group_obj['observedByAll']))

    grp_members = _get_group_members(group_obj['targetName'], group_obj['id'])
    if len(grp_members) == 0 or not set(supervisors).issubset(set(grp_members)):
        not_matching_values.append('members:(cell=[%s],group=[%s])' % (supervisors, grp_members))

    if len(not_matching_values) > 0:
        match = False
        _logger.error('Group "%s" DOES NOT MATCH the source worksheet.%s', group_obj['targetName'],
                      ', '.join(not_matching_values))
    else:
        _logger.info('Group "%s" matches the source worksheet', group_obj['targetName'])
    
    return match

def _process_groups(sites_file: Workbook):
    """Retrieves and processes xMatters Group objects.
        
        Retrieves the Groups from the excel sheet and then attempst to verify
        them in xMatters.  If they don't exist yet, then create them.
        Update the spreadsheet with IDs.
        
        Args:
        sites_file (Workbook): Open sites worksheet
        """
    _logger.info('Processing worksheet for Groups.')
    
    # Get the Sites worksheet
    sites_sheet = sites_file["Sites"];
    
    # Get the Admins worksheet
    admins_sheet = sites_file["Admins"];
    
    # Get the Groups worksheet
    groups_sheet = sites_file["Groups"];
    
    for col in groups_sheet.iter_cols(min_row=2, min_col=column_index_from_string('D'), max_col=column_index_from_string('D')):
        for cell in col:
            _logger.debug('Found cell=%s', cell)
            target_name = cell.value
            site_name = groups_sheet.cell(row=cell.row, column=column_index_from_string('E')).value
            site_id = _get_site_id_from_sites_sheet(sites_sheet, site_name)
            _logger.debug('sites_sheet.site_id=[%s]', site_id)
            supervisors = _get_supervisors_from_admins_sheet(admins_sheet, site_name)
            _logger.debug('admins_sheet.supervisors=[%s]', supervisors)
            if site_id:
                group_obj = _get_group(target_name)
                if group_obj:
                    _logger.info('Processing Group "%s", id=[%s] in the %s environment',
                                 target_name, group_obj['id'],
                                 'Non-Production' if config.non_prod else 'Production')
                    if not _group_match(groups_sheet, cell.row, site_id, supervisors, group_obj):
                        # Update the spreadsheet
                        if config.non_prod:
                            groups_sheet.cell(row=cell.row, column=column_index_from_string('C')).value = group_obj['id']
                        else:
                            groups_sheet.cell(row=cell.row, column=column_index_from_string('B')).value = group_obj['id']
                else:
                    _logger.info('Group "%s" does not exist in the %s environment; adding.',
                                 target_name,
                                 'Non-Production' if config.non_prod else 'Production')
                    group_obj = _add_group(target_name, site_id, site_name, supervisors)
                    if group_obj:
                        members = _add_group_members(target_name, site_id, supervisors)
                        if len(members) > 0:
                            if config.non_prod:
                                groups_sheet.cell(row=cell.row, column=column_index_from_string('C')).value = group_obj['id']
                            else:
                                groups_sheet.cell(row=cell.row, column=column_index_from_string('B')).value = group_obj['id']
            else:
                _logger.error('Unable to find Site "%s" for Group %s.', site_name, target_name)

def process(objects_to_process: list):
    """Verify or create the sites for this instance.

    Read the spreadsheet with Four Seasons property info
    Verify sites, creating missing ones if necessary, caching results
    Verify Admins, creating missing ones if necessary, caching results
    Verify Groups, creating missing ones if necessary

    Args:
        none
    """
    global _logger # pylint: disable=global-statement

    ### Get the current logger
    _logger = np_logger.get_logger()

    # Open the excel sheet
    properties_file = load_workbook(config.properties_filename)
    _logger.debug(properties_file.sheetnames)

    # Process the Site objects based on the spreadsheet
    if 'sites' in objects_to_process:
        _process_sites(properties_file)

    # Save any changes
    properties_file.save(config.properties_filename)

    # Process the Admin objects based on the spreadsheet
    if 'admins' in objects_to_process:
        _process_admins(properties_file)
    
    # Save any changes
    properties_file.save(config.properties_filename)

    # Process the Group objects based on the spreadsheet
    if 'groups' in objects_to_process:
        _process_groups(properties_file)
    
    # Save any changes
    properties_file.save(config.properties_filename)

def main():
    """In case we need to execute the module directly"""
    pass

if __name__ == '__main__':
    main()
